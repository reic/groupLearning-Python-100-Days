from openpyxl import load_workbook
from openpyxl import Workbook
from PIL import Image, ImageFilter


def main1():
    image = Image.open("./guido.jpg")
    # print(image.format, image.size, image.mode)

    # 影像裁切
    rect = 80, 20, 310, 360
    # image.crop(rect).show()
    size = 128, 128
    image.thumbnail(size)
    image.show()


def cut_paste():
    image1 = Image.open("./luohao.png")
    image2 = Image.open("./guido.jpg")
    rect = 80, 20, 310, 360
    guido_head = image2.crop(rect)
    width, height = guido_head.size
    image1.paste(guido_head.resize(
        (int(width/1.5), int(height/1.5))), (172, 40))
    image1.show()


def imageFileter():
    image = Image.open("./guido.jpg")
    image.filter(ImageFilter.CONTOUR).show()


def main():
    workbook = load_workbook('./123.xlsx')
    print(workbook.sheetnames)
    sheet = workbook[workbook.sheetnames[2]]
    print(sheet.title)
    row = sheet.max_row
    print(row)
    print(sheet.cell(row, 1).value)
    # for row in range(2, 3):
    #     for col in range(65, 70):
    #         cell_index = chr(col) + str(row)
    #         print(sheet[cell_index].value, end='\t')
    #     print()
    # print(sheet.max_column, sheet.max_row)
    # print(sheet.cell(3, 2).value)


if __name__ == "__main__":
    main()
